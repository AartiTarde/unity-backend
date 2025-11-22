"""
Excel Generator - Create Excel files with extracted voter data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
from typing import List, Dict
from voter_id_corrector import correct_voter_id


def generate_excel(data: List[Dict], output_path: str) -> bool:
    """
    Generate Excel file from extracted data
    
    Args:
        data: Array of extracted voter data
        output_path: Path to save Excel file
    
    Returns:
        True if successful, raises exception on error
    """
    try:
        # Create workbook and worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Voter Data'
        
        # Define headers
        headers = [
            'EPIC No',
            'Name',
            'Name (English)',
            'Relative Name',
            'Relative Name (English)',
            'Relative Type',
            'House Number',
            'Gender',
            'Age',
            'Assembly Number',
            'Serial Number',
            'Booth Center',
            'Booth Address',
            'Base64 Image String'
        ]
        
        # Set column headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12, color='FFFFFFFF')
            cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 20  # EPIC No
        worksheet.column_dimensions['B'].width = 30  # Name
        worksheet.column_dimensions['C'].width = 30  # Name (English)
        worksheet.column_dimensions['D'].width = 30  # Relative Name
        worksheet.column_dimensions['E'].width = 30  # Relative Name (English)
        worksheet.column_dimensions['F'].width = 15  # Relative Type
        worksheet.column_dimensions['G'].width = 15  # House Number
        worksheet.column_dimensions['H'].width = 10  # Gender
        worksheet.column_dimensions['I'].width = 10  # Age
        worksheet.column_dimensions['J'].width = 20  # Assembly Number
        worksheet.column_dimensions['K'].width = 15  # Serial Number
        worksheet.column_dimensions['L'].width = 25  # Booth Center
        worksheet.column_dimensions['M'].width = 30  # Booth Address
        worksheet.column_dimensions['N'].width = 80  # Base64 Image String
        
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add data rows
        for index, record in enumerate(data):
            row_num = index + 2  # Start from row 2 (row 1 is headers)
            
            # Add voter ID (apply corrections as a safeguard)
            voter_id = record.get('voterID', '')
            if voter_id:
                voter_id = correct_voter_id(voter_id)
            voter_id_cell = worksheet.cell(row=row_num, column=1)
            voter_id_cell.value = voter_id
            voter_id_cell.border = thin_border
            
            # Add name
            name_cell = worksheet.cell(row=row_num, column=2)
            name_cell.value = record.get('name', '')
            name_cell.border = thin_border
            
            # Add name (English)
            name_english_cell = worksheet.cell(row=row_num, column=3)
            name_english_cell.value = record.get('nameEnglish', '')
            name_english_cell.border = thin_border
            
            # Add relative name
            relative_name_cell = worksheet.cell(row=row_num, column=4)
            relative_name_cell.value = record.get('relativeName', '')
            relative_name_cell.border = thin_border
            
            # Add relative name (English)
            relative_name_english_cell = worksheet.cell(row=row_num, column=5)
            relative_name_english_cell.value = record.get('relativeNameEnglish', '')
            relative_name_english_cell.border = thin_border
            
            # Add relative type
            relative_type_cell = worksheet.cell(row=row_num, column=6)
            relative_type_cell.value = record.get('relativeType', '')
            relative_type_cell.border = thin_border
            
            # Add house number
            house_number_cell = worksheet.cell(row=row_num, column=7)
            house_number_cell.value = record.get('houseNumber', '')
            house_number_cell.border = thin_border
            
            # Add gender
            gender_cell = worksheet.cell(row=row_num, column=8)
            gender_cell.value = record.get('gender', '')
            gender_cell.border = thin_border
            
            # Add age
            age_cell = worksheet.cell(row=row_num, column=9)
            age_cell.value = record.get('age', '')
            age_cell.border = thin_border
            
            # Add assembly number
            assembly_number_cell = worksheet.cell(row=row_num, column=10)
            assembly_number_cell.value = record.get('assemblyNumber', '')
            assembly_number_cell.border = thin_border
            
            # Add serial number
            serial_number_cell = worksheet.cell(row=row_num, column=11)
            serial_number_cell.value = record.get('serialNumber', '')
            serial_number_cell.border = thin_border
            
            # Add booth center
            booth_center_cell = worksheet.cell(row=row_num, column=12)
            booth_center_cell.value = record.get('boothCenter', '')
            booth_center_cell.border = thin_border
            
            # Add booth address
            booth_address_cell = worksheet.cell(row=row_num, column=13)
            booth_address_cell.value = record.get('boothAddress', '')
            booth_address_cell.border = thin_border
            
            # Add base64 image
            image_cell = worksheet.cell(row=row_num, column=14)
            image_cell.value = record.get('image_base64', '')
            image_cell.border = thin_border
            
            # Alternate row colors for better readability
            if index % 2 == 0:
                light_gray = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
                voter_id_cell.fill = light_gray
                name_cell.fill = light_gray
                name_english_cell.fill = light_gray
                relative_name_cell.fill = light_gray
                relative_name_english_cell.fill = light_gray
                relative_type_cell.fill = light_gray
                house_number_cell.fill = light_gray
                gender_cell.fill = light_gray
                age_cell.fill = light_gray
                assembly_number_cell.fill = light_gray
                serial_number_cell.fill = light_gray
                booth_center_cell.fill = light_gray
                booth_address_cell.fill = light_gray
                image_cell.fill = light_gray
        
        # Add borders to header row
        for col_num in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.border = thin_border
        
        # Save workbook
        workbook.save(output_path)
        print(f"Excel file generated: {output_path}")
        
        return True
    
    except Exception as e:
        print(f"Excel generation error: {str(e)}")
        raise

